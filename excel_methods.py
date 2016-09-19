from openpyxl import Workbook
from openpyxl import load_workbook
import database_methods
import read_file_methods

def parse_excel():
    wb2 = load_workbook(filename = 'C:\Users\jesong\pp\\file_to_load\\ExcelExport.xlsx', read_only = True, data_only = True)
    name = wb2.get_sheet_names()[0]

    ws = wb2.active
    lst_lst_parsed_file = []
    for row in ws.rows:
    	lst_field = []
    	for cell in row:
            lst_field.append(str(cell.value))
        #print lst_field
        lst_lst_parsed_file.append(lst_field)	

    return lst_lst_parsed_file

if __name__ == '__main__':
    #print parse_excel()
    lst_lst_parsed_file = parse_excel()            
    lst_column_max_length = read_file_methods.get_fields_max_length(lst_lst_parsed_file) 
    lst_field_type = read_file_methods.get_data_type(lst_column_max_length)
    lst_column_name = read_file_methods.add_underscore(lst_lst_parsed_file[0])        
    #table_name = read_file_methods.get_table_name('jeff_excel', str_absolute_path, str_prefix)

    #database_methods.create_table('jeff_excel', lst_column_name, lst_field_type)
    database_methods.insert_rows('jeff_excel', lst_lst_parsed_file, 1)
    



# from openpyxl import load_workbook
# wb = load_workbook(filename='large_file.xlsx', read_only=True)
# ws = wb['big_data'] # ws is now an IterableWorksheet
# for row in ws.rows:
# for cell in row:
# print(cell.value)

            