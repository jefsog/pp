import re
import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

class ClassName(object):
    """docstring for ClassName"""
    def __init__(self, arg):
        super(ClassName, self).__init__()
        self.arg = arg
        
    # use Python CSV library to parse files
    def csv_reader(str_path_name, str_delimiter = ','):
        '''    
        '\0' is null bytes in the input file    
        '''
        
        lst_lst_field = []
        try:
            if str_delimiter == ',':
                
                with open(str_path_name, 'rb') as csvfile:
                    reader = csv.reader(csvfile)            
                    for row in reader:
                        lst_lst_field.append(row)
                
            else:
                
                with open(str_path_name, 'rb') as csvfile:
                    reader = csv.reader(csvfile,  delimiter=str_delimiter, quoting=csv.QUOTE_NONE)            
                    for row in reader:
                        #print row
                        lst_lst_field.append(row)        
        
        except csv.Error as e:
            print str_path_name + ':' + e
        return lst_lst_field


    #return a list of fields from a single line
    #can only deal with fields delimited by comma and text wrapped by '"'
    def parse_line(str_line): 
        #lst_fields = re.findall(r'.*?(,|",)',str_line)
        lst_fields = re.findall(r'(".*?(?:",|"$))', str_line)
        
        #lst_fields = re.findall(r'(".*?")', str_line)
        
        for i in range(len(lst_fields)):
            if i < len(lst_fields)-1:
                lst_fields[i] = lst_fields[i][1:-2]
            else:
                lst_fields[i] = lst_fields[i][1:-1]
        
        return lst_fields

    # deal with excel style csv
    def parse_line_excel_csv(str_line): 
        lst_position = pop_extra_comma(get_character_position(',', str_line), pop_adjacent_numbers(get_character_position('"', str_line)))    
        lst_fields = []
        lst_fields.append(clean_field(str_line[0:lst_position[0]]))
        for i in range(len(lst_position)-1):
            lst_fields.append(clean_field(str_line[lst_position[i]:lst_position[i+1]]))
        
        lst_fields.append(clean_field(str_line[lst_position[len(lst_position)-1]:len(str_line)-1]))
        
        return lst_fields    


    def get_character_position(str_character, str_line):
        lst_comma = []
        for i in range(len(str_line)):
            if str_line[i] == str_character:
                lst_comma.append(i)
            
        return lst_comma

    def pop_adjacent_numbers(lst_number):
        lst_number_no_adjacent = []
        for i in range(len(lst_number)):
            if(i != len(lst_number)-1 and lst_number[i]+1 == lst_number[i+1] ):
                i = i+2
                
            else:
                lst_number_no_adjacent.append(lst_number[i])
        return lst_number_no_adjacent

    def pop_extra_comma(lst_comma, lst_quotation):
        lst_comma_extra_removed = copy_list(lst_comma)
        i = len(lst_comma)-1
        while i>=0:
            j = 0
            while j < len(lst_quotation)-1:
                if lst_comma[i]>lst_quotation[j] and lst_comma[i]<lst_quotation[j+1]:
                    lst_comma_extra_removed.pop(i)
                #print [lst_quotation[j], lst_quotation[j+1]]
                
                j = j+2
            
            
            i = i-1
        return lst_comma_extra_removed

    def copy_list(lst_list):
        lst_cloan = []
        for item in lst_list:
            lst_cloan.append(item)
        return lst_cloan

    def clean_field(str_field):
        if str_field[0] == ',':
            str_field = str_field[1:]
        if len(str_field) > 1 and str_field[0] == '"' and str_field[-1] == '"':
            str_field = str_field[1:-1]
        str_field = str_field.replace('""', '"')
        return str_field

    #read a txt file, return a list of lines in the file
    def read_file(str_file_name): 
        file_in = open(str_file_name, 'rU')
        lst_line = []
        for line in file_in:
            lst_line.append(line)
        file_in.close()
        return lst_line

    #return a list of max length of each column in the parsed file
    def get_fields_max_length(lst_lst_file): 
        lst_field_max_length = []
        for j in range(1, len(lst_lst_file)): #skip the file head, start from the second line
            line = lst_lst_file[j]
            i = 0
            for field in line:
                if len(lst_field_max_length) <= i:
                    lst_field_max_length.insert(i, -1) #fill the empty list with -1 in order to compare 
                if len(field) > lst_field_max_length[i]:
                    lst_field_max_length[i] = len(field)
                i = i+1
        
        return lst_field_max_length

    #return a list of absolute path name in a directory
    def read_directory(str_absolute_path):
        lst_file_name = os.listdir(str_absolute_path)
        lst_absolute_path_name = []
        for name in lst_file_name:

            if name[0:8] != 'load_log':
                lst_absolute_path_name.append(os.path.join(str_absolute_path, name)) 
        return lst_absolute_path_name


    def get_data_type(lst_field_length):
        lst_data_type = []
        for num in lst_field_length:
            if num<=4000:
                if num == 0:  #zero length varchar is not allowed
                    num = 1; 
                lst_data_type.append('varchar('+str(num)+')')
            else:
                lst_data_type.append('clob')
                
        return lst_data_type






    # if the column is delimited by space, replace the space with underscore
    def add_underscore(lst_string):    
        lst_underscore_added = []
        for content in lst_string:
            content = content.strip()
            lst_underscore_added.append(content.replace(' ', '_'))
        return lst_underscore_added

    # retrieve file name, prefix it with a string, trunk its length to no more than 30
    def get_table_name(str_absolute_file_name, str_directory, str_prefix):
        int_dot_position = str_absolute_file_name.find('.')
        table_name = str_prefix + '_' + str_absolute_file_name[len(str_directory)+1:int_dot_position]
        
        while table_name.find('  ') > -1:
            table_name = table_name.replace('  ', ' ')
        if table_name.find(' ') > -1:
            table_name = table_name.replace(' ', '_')
        if table_name.find('-') > -1:
            table_name = table_name.replace('-', '_')
        while table_name.find('__') > -1:
            table_name = table_name.replace('__', '_')    
        if len(table_name) > 30:
            table_name = table_name[0:29]
        return table_name


class CSV_reader(File_reader):
    """read csv or txt file, return a list of rows and fields"""
    def __init__(self, arg):
        super(CSV_reader, self).__init__()
        self.arg = arg
        
    # use Python CSV library to parse files
    def read(str_path_name, str_delimiter = ','):
        '''    
        '\0' is null bytes in the input file    
        '''
        
        lst_lst_field = []
        try:
            if str_delimiter == ',':
                
                with open(str_path_name, 'rb') as csvfile:
                    reader = csv.reader(csvfile)            
                    for row in reader:
                        lst_lst_field.append(row)
                
            else:
                
                with open(str_path_name, 'rb') as csvfile:
                    reader = csv.reader(csvfile,  delimiter=str_delimiter, quoting=csv.QUOTE_NONE)            
                    for row in reader:
                        #print row
                        lst_lst_field.append(row)        
        
        except csv.Error as e:
            print str_path_name + ':' + e
        return lst_lst_field



def parse_excel():
    wb2 = load_workbook(filename = 'C:\Users\jesong\pp\\file_to_load\\ExcelExport.xlsx', read_only = True, data_only = True)
    name = wb2.get_sheet_names()[0]

    ws = wb2.active
    lst_lst_parsed_file = []
    for row in ws.rows:
        lst_field = []
        for cell in row:
            lst_field.append(str(cell.value))
        #print lst_field
        lst_lst_parsed_file.append(lst_field)   

    return lst_lst_parsed_file

if __name__ == '__main__':
    #print parse_excel()
    lst_lst_parsed_file = parse_excel()            
    lst_column_max_length = read_file_methods.get_fields_max_length(lst_lst_parsed_file) 
    lst_field_type = read_file_methods.get_data_type(lst_column_max_length)
    lst_column_name = read_file_methods.add_underscore(lst_lst_parsed_file[0])        
    #table_name = read_file_methods.get_table_name('jeff_excel', str_absolute_path, str_prefix)

    #database_methods.create_table('jeff_excel', lst_column_name, lst_field_type)
    database_methods.insert_rows('jeff_excel', lst_lst_parsed_file, 1)
    



# from openpyxl import load_workbook
# wb = load_workbook(filename='large_file.xlsx', read_only=True)
# ws = wb['big_data'] # ws is now an IterableWorksheet
# for row in ws.rows:
# for cell in row:
# print(cell.value)

            